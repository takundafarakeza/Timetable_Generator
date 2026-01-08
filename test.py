from utils.builders import PrimaryBuilder, SecondaryBuilder, TertiaryBuilder
from PySide6.QtWidgets import QApplication
from utils import Types
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json
import os
import sys

file = open(
    r"C:\Users\TAKUNDA\AppData\Roaming\Timetable_Generator\timetables\First_Term_Timetable_Belfair_Secondary.tbl", "r")
file_data = json.loads(file.read())
file.close()

slots_per_day = 12
days_per_cycle = 5

header_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

work_book = Workbook()
work_book.remove(work_book.active)


class Exc:
    def __init__(self):
        self.builder = TertiaryBuilder(r"C:\Users\TAKUNDA\AppData\Roaming\Timetable_Generator"
                                      r"\timetables\Timetable-Nest_Univertsity.tbl")

    def export(self):
        timetable_data = self.builder.timetable_get()
        export_data = {}
        for day in timetable_data:
            export_data[day] = {}
            for slot_i, slot in enumerate(timetable_data[day]):
                export_data[day][slot_i + 2] = {}
                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                    if isinstance(timetable_data[day][slot], dict):
                        for col, event in enumerate(timetable_data[day][slot][Types.EVENTS]):
                            if timetable_data[day][slot][Types.EVENTS][event][Types.SUBJECT][0] == "b":
                                block = self.builder.block_get(timetable_data[day][slot][Types.EVENTS]
                                                               [event][Types.SUBJECT])
                                subjects = ""
                                for subject in block.subjects:
                                    subjects += (f"({self.builder.subject_get(subject).name}:   "
                                                 f"{self.builder.teacher_get(block.subjects[subject]
                                                                             [Types.TEACHER]).name}: "
                                                 f"{self.builder.venue_get(block.subjects[subject]
                                                                           [Types.VENUE]).name})")
                                export_data[day][
                                    slot_i + 2][col + 2] = (f"{self.builder.class_get(event).name}:   "
                                                            f"{block.name}\n {subjects}")

                            else:
                                export_data[day][slot_i + 2][col + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                         f"{self.builder.subject_get(timetable_data[day]
                                                                                                     [slot]
                                                                                                     [Types.EVENTS]
                                                                                                     [event][Types.
                                                                                                     SUBJECT]).name}: "
                                                                         f"{self.builder.teacher_get(timetable_data
                                                                                                     [day][slot]
                                                                                                     [Types.EVENTS]
                                                                                                     [event]
                                                                                                     [Types.TEACHER]
                                                                                                     ).name}")
                    else:
                        export_data[day][slot_i + 2][2] = timetable_data[day][slot]
                else:
                    if isinstance(timetable_data[day][slot], dict):
                        for col, event in enumerate(timetable_data[day][slot]):
                            if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                export_data[day][slot_i + 2][col + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                         f"{self.builder.subject_get(timetable_data[day]
                                                                                                     [slot][event]
                                                                                                     [Types.SUBJECT]
                                                                                                     ).name}")
                            elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                venue = self.builder.venue_get(timetable_data[day][slot][event][Types.VENUE]).name
                                courses = ", ".join(timetable_data[day][slot][event][Types.COURSES])
                                export_data[day][slot_i + 2][col + 2] = (f"{self.builder.module_get(event).name}:   "
                                                                         f"{courses}: {venue}")
                    else:
                        export_data[day][slot_i + 2][2] = timetable_data[day][slot]

        vertical_header = []
        slots = self.builder.timetable_slots()

        for slot in range(self.builder.timetable_get_slots_per_day()):
            slot = str(slot + 1)
            slot_item = slot if slot not in slots else slots[slot]
            vertical_header.append(slot_item)

        for day in export_data:
            days = self.builder.timetable_days()
            work_sheet = work_book.create_sheet(days[day] if day in days else f"Day {day}")
            work_sheet.append(["Period"])
            for row in range(2, len(vertical_header) + 2):
                work_sheet.append([slots[str(row - 1)]] if str(row - 1) in slots else [str(row - 1)])
                cell = work_sheet.cell(row=row, column=1)
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            for slot in export_data[day]:
                for col in export_data[day][slot]:
                    work_sheet.cell(row=slot, column=col, value=export_data[day][slot][col])
                    app.quit()

    def export_filtered(self):
        filter_ = "2"  # self.filter.currentData()
        timetable_data = self.builder.timetable_get()
        export_data = {}
        for row, day in enumerate(timetable_data):
            export_data[day] = {}
            for slot_i, slot in enumerate(timetable_data[day]):
                export_data[day][slot_i + 2] = {}
                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                    if isinstance(timetable_data[day][slot], dict):
                        for event in timetable_data[day][slot][Types.EVENTS]:
                            if event == filter_:
                                if timetable_data[day][slot][Types.EVENTS][event][Types.SUBJECT][0] == "b":
                                    block = self.builder.block_get(timetable_data[day][slot][Types.EVENTS]
                                                                   [event][Types.SUBJECT])
                                    subjects = ""
                                    for subject in block.subjects:
                                        subjects += (f"({self.builder.subject_get(subject).name}:   "
                                                     f"{self.builder.teacher_get(block.subjects[subject]
                                                                                 [Types.TEACHER]).name}: "
                                                     f"{self.builder.venue_get(block.subjects[subject]
                                                                               [Types.VENUE]).name})")
                                    export_data[day][
                                        slot_i + 2][row + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                f"{block.name}\n {subjects}")

                                else:
                                    export_data[day][slot_i + 2][row + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                             f"{self.builder.subject_get(timetable_data[day]
                                                                                                         [slot]
                                                                                                         [Types.EVENTS]
                                                                                                         [event][Types.
                                                                                                         SUBJECT]).name}: "
                                                                             f"{self.builder.teacher_get(timetable_data
                                                                                                         [day][slot]
                                                                                                         [Types.EVENTS]
                                                                                                         [event]
                                                                                                         [Types.TEACHER]
                                                                                                         ).name}")
                    else:
                        export_data[day][slot_i + 2][day + 1] = timetable_data[day][slot]
                else:
                    if isinstance(timetable_data[day][slot], dict):
                        for event in timetable_data[day][slot]:
                            if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                if filter_ == event:
                                    export_data[day][slot_i + 2][row + 2] = (f"{self.builder.subject_get(timetable_data
                                                                                                         [day]
                                                                                                         [slot][event]
                                                                                                         [Types.SUBJECT]
                                                                                                         ).name}")
                            elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                module = self.builder.module_get(event)

                                if filter_ in module.courses:
                                    venue = self.builder.venue_get(timetable_data[day][slot][event][Types.VENUE]).name
                                    courses = ", ".join(timetable_data[day][slot][event][Types.COURSES])
                                    export_data[day][slot_i + 2][row + 2] = (
                                        f"{self.builder.module_get(event).name}:   "
                                        f"{courses}: {venue}")
                    else:
                        export_data[day][slot_i + 2][row + 2] = timetable_data[day][slot]

        horizontal_header = [""]

        slots = self.builder.timetable_slots()
        for slot in range(self.builder.timetable_get_slots_per_day()):
            slot = str(slot + 1)
            slot_item = slot if slot not in slots else slots[slot]
            horizontal_header.append(slot_item)

        work_sheet = work_book.create_sheet(self.builder.timetable_get_name())
        work_sheet.append(horizontal_header)

        for slot_i in range(2, len(horizontal_header) + 2):
            cell = work_sheet.cell(row=1, column=slot_i)
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        days = self.builder.timetable_days()

        for day in export_data:
            row = int(day) + 1
            cell = work_sheet.cell(row=row, column=1, value=(days[str(row - 1)] if str(row - 1) in days else
                                                             f"Day {str(row - 1)}"))
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            for slot in export_data[day]:
                for row in export_data[day][slot]:
                    work_sheet.cell(row=row, column=slot, value=export_data[day][slot][row])
                    app.quit()


app = QApplication()
Exc().export_filtered()
work_book.save("timetable.xlsx")
os.startfile("timetable.xlsx")
sys.exit(app.exec())
