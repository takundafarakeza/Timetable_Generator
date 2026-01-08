from PySide6.QtWidgets import (QMainWindow, QMessageBox,
                               QGraphicsDropShadowEffect, QPushButton,
                               QMenu, QWidget, QDialog, QFileDialog)
from utils.builders import PrimaryBuilder, SecondaryBuilder, TertiaryBuilder
from PySide6.QtGui import QGuiApplication, QColor, QIcon
from PySide6.QtCore import Qt, Signal, QSize
from ui import ui_startup, ui_boarding, ui_main_window, ui_viewer, ui_export_project_dialog
from utils import Types, Settings
from utils.logger_config import logger
from typing import Optional, Union
from widgets import RecentItem, TimeTable, MessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import functools


class Viewer(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = ui_viewer.Ui_ViewerWindow()
        self.ui.setupUi(self)
        self.builder: Optional[PrimaryBuilder, SecondaryBuilder, TertiaryBuilder] = None
        self.tab_view = self.ui.tab_view
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.filter = self.ui.filter

        self.ui.close_btn.clicked.connect(self.close)
        self.ui.minimize_btn.clicked.connect(self.showMinimized)
        self.ui.maximize_btn.clicked.connect(self.showMaximized)

    def showMaximized(self, /):
        if self.isMaximized():
            self.showNormal()
        else:
            super().showMaximized()

    def clear(self):
        self.filter.currentTextChanged.disconnect()
        self.tab_view.clear()

    def close(self, /):
        self.clear()
        super().close()

    def view(self, builder):
        self.builder = builder
        if self.builder.timetable_get():
            self.load()
        self.show()

    def load(self):
        try:
            self.tab_view.clear()
            self.filter.clear()
            timetable_data = self.builder.timetable_get()
            filter_data = []

            for day in timetable_data:
                timetable = TimeTable(self.builder.timetable_get_slots_per_day(), self.builder.timetable_slots())
                self.tab_view.addTab(timetable, (f"Day {day}" if day not in self.builder.timetable_days() else
                                                 self.builder.timetable_days()[day]))
                for i, slot in enumerate(timetable_data[day]):
                    if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot][Types.EVENTS]:
                                if timetable_data[day][slot][Types.EVENTS][event][Types.SUBJECT][0] == "b":
                                    block = self.builder.block_get(timetable_data[day][slot][Types.EVENTS]
                                                                   [event][Types.SUBJECT])
                                    subjects = ((f"({self.builder.subject_get(subject).name}: "
                                                 f"{self.builder.teacher_get(block.subjects[subject]
                                                                             [Types.TEACHER]).name}: "
                                                 f"{self.builder.venue_get(block.subjects[subject]
                                                                           [Types.VENUE]).name})")
                                                for subject in block.subjects)
                                    subjects = ", ".join(subjects)
                                    filter_data.append((event, self.builder.class_get(event).name))
                                    timetable.add_item(i, self.builder.class_get(event).name, block.name, subjects)
                                else:
                                    timetable.add_item(i, self.builder.class_get(event).name,
                                                       self.builder.subject_get(timetable_data
                                                                                [day][slot][Types.EVENTS][event]
                                                                                [Types.SUBJECT]).name,
                                                       self.builder.teacher_get(timetable_data
                                                                                [day][slot][Types.EVENTS][event]
                                                                                [Types.TEACHER]).name)
                                    filter_data.append((event, self.builder.class_get(event).name))
                        else:
                            timetable.add_item(i, timetable_data[day][slot], "", "")
                    else:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot]:
                                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                    timetable.add_item(i, self.builder.class_get(event).name,
                                                       self.builder.subject_get(timetable_data
                                                                                [day][slot][event]
                                                                                [Types.SUBJECT]).name, "")
                                    filter_data.append((event, self.builder.class_get(event).name))

                                elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                    venue = self.builder.venue_get(timetable_data[day][slot][event][Types.VENUE]).name
                                    courses = ", ".join(timetable_data[day][slot][event][Types.COURSES])
                                    timetable.add_item(i, self.builder.module_get(event).name, courses, venue)
                                    filter_data.append((event, self.builder.module_get(event).name))
                        else:
                            timetable.add_item(i, timetable_data[day][slot], "", "")

            filter_data = {filter_item for filter_item in filter_data}
            self.filter.addItem("All")
            for filter_item in filter_data:
                self.filter.addItem(filter_item[1], userData=filter_item[0])
            self.filter.setCurrentText("All")
            self.filter.currentTextChanged.connect(self.load_filtered)
        except Exception as e:
            logger.critical(str(e))

    def load_filtered(self):

        if self.filter.currentText() == "All":
            self.filter.currentTextChanged.disconnect()
            self.load()
            return

        try:
            self.tab_view.clear()
            timetable_data = self.builder.timetable_get()

            for day in timetable_data:
                timetable = TimeTable(self.builder.timetable_get_slots_per_day(), self.builder.timetable_slots())
                self.tab_view.addTab(timetable,
                                     f"Day {day}" if day not in self.builder.timetable_days() else self.builder.
                                     timetable_days()[day])
                for i, slot in enumerate(timetable_data[day]):
                    if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                        for event in timetable_data[day][slot][Types.EVENTS]:
                            if isinstance(timetable_data[day][slot][Types.EVENTS], dict):
                                if event == self.filter.currentData():
                                    if timetable_data[day][slot][Types.EVENTS][event][Types.SUBJECT][0] == "b":
                                        block = self.builder.block_get(timetable_data[day][slot][Types.EVENTS]
                                                                       [event][Types.SUBJECT])
                                        subjects = (f"({self.builder.subject_get(subject).name}: "
                                                    f"{self.builder.teacher_get(block.subjects[subject]
                                                                                [Types.TEACHER]).name})"
                                                    for subject in block.subjects)
                                        subjects = ", ".join(subjects)
                                        timetable.add_item(i, self.builder.class_get(event).name, block.name, subjects)
                                    else:
                                        timetable.add_item(i, self.builder.class_get(event).name,
                                                           self.builder.subject_get(timetable_data
                                                                                    [day][slot][Types.EVENTS][event]
                                                                                    [Types.SUBJECT]).name,
                                                           self.builder.teacher_get(timetable_data
                                                                                    [day][slot][Types.EVENTS][event]
                                                                                    [Types.TEACHER]).name)
                            else:
                                timetable.add_item(i, timetable_data[day][slot][Types.EVENTS], "", "")
                    else:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot]:
                                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                    if event == self.filter.currentData():
                                        timetable.add_item(i, self.builder.class_get(event).name,
                                                           self.builder.subject_get(timetable_data
                                                                                    [day][slot][event]
                                                                                    [Types.SUBJECT]).name, "")

                                elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                    if event == self.filter.currentData():
                                        venue = self.builder.venue_get(
                                            timetable_data[day][slot][event][Types.VENUE]).name
                                        courses = ", ".join(timetable_data[day][slot][event][Types.COURSES])
                                        timetable.add_item(i, self.builder.module_get(event).name, courses, venue)
                        else:
                            timetable.add_item(i, timetable_data[day][slot], "", "")
        except Exception as e:
            logger.critical(str(e))


class Export(QDialog):
    def __init__(self, builder: Union[PrimaryBuilder, SecondaryBuilder, TertiaryBuilder], parent):
        super().__init__(parent)
        self.ui = ui_export_project_dialog.Ui_Export()
        self.ui.setupUi(self)
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.FramelessWindowHint)
        self.setWindowModality(Qt.WindowModality.WindowModal)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)

        self.builder = builder

        self.ui.close_btn.clicked.connect(self.close)
        self.ui.cancel_btn.clicked.connect(self.close)
        self.filter = self.ui.timetable_filter
        self.ui.timetable_export.clicked.connect(self.export)
        self.populate_filter()

        self.header_font = Font(bold=True)
        self.center = Alignment(horizontal="center", vertical="center")
        self.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )

    def populate_filter(self):
        try:
            self.filter.clear()
            timetable_data = self.builder.timetable_get()
            filter_data = []

            for day in timetable_data:
                for i, slot in enumerate(timetable_data[day]):
                    if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot][Types.EVENTS]:
                                filter_data.append((event, self.builder.class_get(event).name))
                    else:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot]:
                                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                    filter_data.append((event, self.builder.class_get(event).name))
                                elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                    for course in self.builder.module_get(event).courses:
                                        filter_data.append((course, self.builder.course_get(course).name))

            filter_data = {filter_item for filter_item in filter_data}
            self.filter.addItem("All", userData="all")
            for filter_item in filter_data:
                self.filter.addItem(filter_item[1], userData=filter_item[0])
            self.filter.setCurrentText("All")
        except Exception as e:
            logger.critical(str(e))

    def export(self):

        if self.filter.currentData() != "all":
            self.export_filtered()
            return

        try:
            timetable_data = self.builder.timetable_get()
            work_book = Workbook()
            work_book.remove(work_book.active)
            export_data = {}

            for day in timetable_data:
                export_data[day] = {}
                for slot_i, slot in enumerate(timetable_data[day]):
                    export_data[day][slot_i + 2] = {}
                    if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                        if isinstance(timetable_data[day][slot], dict):
                            for col, event in enumerate(timetable_data[day][slot][Types.EVENTS]):
                                if timetable_data[day][slot][Types.EVENTS][event][Types.SUBJECT][0] == "b":
                                    block = self.builder.block_get(timetable_data[day][slot][Types.EVENTS]
                                                                   [event][Types.SUBJECT])
                                    subjects = ""
                                    for subject in block.subjects:
                                        subjects += (f"({self.builder.subject_get(subject).name}:   "
                                                     f"{self.builder.teacher_get(block.subjects[subject]
                                                                                 [Types.TEACHER]).name}: "
                                                     f"{self.builder.venue_get(block.subjects[subject]
                                                                               [Types.VENUE]).name})")
                                    export_data[day][
                                        slot_i + 2][col + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                f"{block.name}\n {subjects}")

                                else:
                                    export_data[day][slot_i + 2][col + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                             f"{self.builder.subject_get(
                                                                                 timetable_data[day][slot][Types.EVENTS]
                                                                                 [event][Types.SUBJECT]).name}: "
                                                                             f"{self.builder.teacher_get(
                                                                                 timetable_data[day][slot][Types.EVENTS]
                                                                                 [event][Types.TEACHER]).name}")
                        else:
                            export_data[day][slot_i + 2][2] = timetable_data[day][slot]
                    else:
                        if isinstance(timetable_data[day][slot], dict):
                            for col, event in enumerate(timetable_data[day][slot]):
                                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                    export_data[day][slot_i + 2][col + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                             f"{self.builder.subject_get(
                                                                                 timetable_data[day][slot][event]
                                                                                 [Types.SUBJECT]).name}")
                                elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                    venue = self.builder.venue_get(timetable_data[day][slot][event][Types.VENUE]).name
                                    courses = ", ".join(timetable_data[day][slot][event][Types.COURSES])
                                    export_data[day][slot_i + 2][col + 2] = (f"{self.builder.module_get(event).name}"
                                                                             f":   {courses}: {venue}")
                        else:
                            export_data[day][slot_i + 2][2] = timetable_data[day][slot]

            vertical_header = []
            slots = self.builder.timetable_slots()

            for slot in range(self.builder.timetable_get_slots_per_day()):
                slot = str(slot + 1)
                slot_item = slot if slot not in slots else slots[slot]
                vertical_header.append(slot_item)

            for day in export_data:
                days = self.builder.timetable_days()
                work_sheet = work_book.create_sheet(days[day] if day in days else f"Day {day}")
                work_sheet.append(["Period"])
                cell = work_sheet.cell(row=1, column=1)
                cell.font = self.header_font
                cell.alignment = self.center
                cell.border = self.border

                for row in range(2, len(vertical_header) + 2):
                    work_sheet.append([slots[str(row - 1)]] if str(row - 1) in slots else [str(row - 1)])
                    cell = work_sheet.cell(row=row, column=1)
                    cell.font = self.header_font
                    cell.alignment = self.center
                    cell.border = self.border
                for slot in export_data[day]:
                    for col in export_data[day][slot]:
                        work_sheet.cell(row=slot, column=col, value=export_data[day][slot][col])

            save_name = self.builder.timetable_get_name().replace(" ", "_") + ".xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Timetable", save_name, "Excel Files (*.xlsx)"
            )
            if file_path:
                work_book.save(file_path)

        except Exception as e:
            logger.critical(str(e))
            MessageBox(self).critical("Error", "This is embarrassing! An unexpected error occurred.")

    def export_filtered(self):

        try:
            filter_ = self.filter.currentData()
            timetable_data = self.builder.timetable_get()
            work_book = Workbook()
            work_book.remove(work_book.active)
            export_data = {}

            for row, day in enumerate(timetable_data):
                export_data[day] = {}
                for slot_i, slot in enumerate(timetable_data[day]):
                    export_data[day][slot_i + 2] = {}
                    if self.builder.timetable_get_institution_type() == Types.INSTITUTION_SECONDARY:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot][Types.EVENTS]:
                                if event == filter_:
                                    if timetable_data[day][slot][Types.EVENTS][event][Types.SUBJECT][0] == "b":
                                        block = self.builder.block_get(timetable_data[day][slot][Types.EVENTS]
                                                                       [event][Types.SUBJECT])
                                        subjects = ""
                                        for subject in block.subjects:
                                            subjects += (f"({self.builder.subject_get(subject).name}:   "
                                                         f"{self.builder.teacher_get(block.subjects[subject]
                                                                                     [Types.TEACHER]).name}: "
                                                         f"{self.builder.venue_get(block.subjects[subject]
                                                                                   [Types.VENUE]).name})")
                                        export_data[day][
                                            slot_i + 2][row + 2] = (f"{self.builder.class_get(event).name}:   "
                                                                    f"{block.name}\n {subjects}")

                                    else:
                                        export_data[day][slot_i + 2][row + 2] = (f"{self.builder.class_get(event).name}"
                                                                                 f":   "
                                                                                 f"{self.builder.subject_get(
                                                                                     timetable_data[day][slot]
                                                                                     [Types.EVENTS][event]
                                                                                     [Types.SUBJECT]).name}: "
                                                                                 f"{self.builder.teacher_get(
                                                                                     timetable_data[day][slot]
                                                                                     [Types.EVENTS][event]
                                                                                     [Types.TEACHER]).name}")
                        else:
                            export_data[day][slot_i + 2][row + 2] = timetable_data[day][slot]
                    else:
                        if isinstance(timetable_data[day][slot], dict):
                            for event in timetable_data[day][slot]:
                                if self.builder.timetable_get_institution_type() == Types.INSTITUTION_PRIMARY:
                                    if filter_ == event:
                                        export_data[day][slot_i + 2][row + 2] = (f"{self.builder.subject_get(
                                            timetable_data[day][slot][event][Types.SUBJECT]).name}")
                                elif self.builder.timetable_get_institution_type() == Types.INSTITUTION_COLLEGE:
                                    module = self.builder.module_get(event)

                                    if filter_ in module.courses:
                                        venue = self.builder.venue_get(timetable_data[day][slot]
                                                                       [event][Types.VENUE]).name
                                        courses = ", ".join(timetable_data[day][slot][event][Types.COURSES])
                                        export_data[day][slot_i + 2][row + 2] = (
                                            f"{self.builder.module_get(event).name}:   "
                                            f"{courses}: {venue}")
                        else:
                            export_data[day][slot_i + 2][row + 2] = timetable_data[day][slot]

            horizontal_header = [""]

            slots = self.builder.timetable_slots()
            for slot in range(self.builder.timetable_get_slots_per_day()):
                slot = str(slot + 1)
                slot_item = slot if slot not in slots else slots[slot]
                horizontal_header.append(slot_item)

            work_sheet = work_book.create_sheet(self.builder.timetable_get_name())
            work_sheet.append(horizontal_header)

            for slot_i in range(2, len(horizontal_header) + 2):
                cell = work_sheet.cell(row=1, column=slot_i)
                cell.font = self.header_font
                cell.alignment = self.center
                cell.border = self.border

            days = self.builder.timetable_days()

            for day in export_data:
                row = int(day) + 1
                cell = work_sheet.cell(row=row, column=1, value=(days[str(row - 1)] if str(row - 1) in days else
                                                                 f"Day {str(row - 1)}"))
                cell.font = self.header_font
                cell.alignment = self.center
                cell.border = self.border
                for slot in export_data[day]:
                    for row in export_data[day][slot]:
                        work_sheet.cell(row=row, column=slot, value=export_data[day][slot][row])

            save_name = self.builder.timetable_get_name().replace(" ", "_") + ".xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Timetable", save_name, "Excel Files (*.xlsx)"
            )
            if file_path:
                work_book.save(file_path)

        except Exception as e:
            logger.critical(str(e))
            MessageBox(self).critical("Error", "This is embarrassing! An unexpected error occurred.")


class Main(QMainWindow):
    file_signal = Signal(str)

    def __init__(self):
        super().__init__()
        self.ui = ui_main_window.Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.geometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()

        self.viewer = Viewer()

        window_width = int(screen_width * 0.85)
        window_height = int(screen_height * 0.85)
        self.ui.setupUi(self)
        self.resize(window_width, window_height)
        self.move((screen_width - window_width) // 2,
                  (screen_height - window_height) // 2)

        self.home_btn = self.ui.home_btn
        self.subjects_btn = self.ui.subjects_btn
        self.modules_btn = self.ui.modules_btn
        self.teachers_btn = self.ui.teachers_btn
        self.lecturers_btn = self.ui.lecturers_btn
        self.classes_btn = self.ui.classes_btn
        self.courses_btn = self.ui.courses_btn
        self.venues_btn = self.ui.venues_btn
        self.blocks_btn = self.ui.blocks_btn
        self.menu_btns = {
            self.home_btn: {"icons": [u":/icons/icons/home.svg", u":/icons/icons/home-white.svg"],
                            "page": 1},
            self.subjects_btn: {"icons": [u":/icons/icons/subject.svg", u":/icons/icons/subject-white.svg"],
                                "page": 3},
            self.modules_btn: {"icons": [u":/icons/icons/subject.svg", u":/icons/icons/subject-white.svg"],
                               "page": 2},
            self.blocks_btn: {"icons": [u":/icons/icons/blocks.svg", u":/icons/icons/blocks-white.svg"],
                              "page": 4},
            self.teachers_btn: {"icons": [u":/icons/icons/teacher.svg", u":/icons/icons/teacher-white.svg"],
                                "page": 6},
            self.lecturers_btn: {"icons": [u":/icons/icons/teacher.svg", u":/icons/icons/teacher-white.svg"],
                                 "page": 5},
            self.classes_btn: {"icons": [u":/icons/icons/classes.svg", u":/icons/icons/classes-white.svg"],
                               "page": 8},
            self.courses_btn: {"icons": [u":/icons/icons/classes.svg", u":/icons/icons/classes-white.svg"],
                               "page": 7},
            self.venues_btn: {"icons": [u":/icons/icons/venues.svg", u":/icons/icons/venues-white.svg"],
                              "page": 9}
        }
        self.display_cards = [self.ui.courses_card, self.ui.venues_card, self.ui.subjects_card,
                              self.ui.modules_card, self.ui.classes_card, self.ui.lecturers_card,
                              self.ui.teachers_card, self.ui.blocks_card]

        self.settings_btn = self.ui.settings_btn
        self.close_project_btn = self.ui.close_project_btn

        self.close_btn = self.ui.close_btn
        self.minimize_btn = self.ui.minimize_btn
        self.maximize_btn = self.ui.maximize_btn

        self.menu_btn = self.ui.menu_btn
        self.file_btn = self.ui.file_btn
        self.tools_btn = self.ui.tools_btn
        self.view_btn = self.ui.view_btn
        self.edit_btn = self.ui.edit_btn
        self.timetable_btn = self.ui.timetable_btn
        self.save_btn = self.ui.save_btn
        self.save_indicator = self.ui.save_indicator
        self.generate_indicator = self.ui.generate_indicator
        self.current_project = self.ui.current_project

        self.minimize_btn.clicked.connect(self.showMinimized)
        self.maximize_btn.clicked.connect(self.showMaximized)

        self.edit_menu = QMenu()
        self.file_menu = QMenu()
        self.tools_menu = QMenu()

        self.view_menu = QMenu()
        self.view_menu.addAction("Fullscreen", self.showFullScreen)
        self.view_menu.addAction("Maximize", self.showMaximized)
        self.view_menu.addAction("Minimize", self.showMinimized)
        self.view_btn.setMenu(self.view_menu)

        self.timetable_menu = QMenu()

        self.slots_table = self.ui.slot_edit_table
        self.days_table = self.ui.day_edit_table
        self.slots_table.setColumnWidth(0, 20)
        self.slots_table.setColumnWidth(1, 160)
        self.days_table.setColumnWidth(0, 20)
        self.days_table.setColumnWidth(1, 160)

        self.setup_menu()

    """
    ======================================================
            UI Manipulation
            20/12/2025
    ======================================================
    """

    def showMaximized(self, /):
        if self.isMaximized():
            self.showNormal()
        else:
            super().showMaximized()

    def showFullScreen(self, /):
        if self.isFullScreen():
            self.showNormal()
        else:
            super().showFullScreen()

    def activate_btn(self, button_: QPushButton):
        icon = QIcon()
        icon.addFile(self.menu_btns[button_]["icons"][1],
                     QSize(), QIcon.Mode.Normal, QIcon.State.Off)
        button_.setIcon(icon)
        button_.setStyleSheet(
            u"background: #2F69B2;"
            "color: #F3F4F6;"
        )
        self.ui.stacked_container.setCurrentIndex(self.menu_btns[button_]["page"] - 1)
        for button__ in self.menu_btns:
            if button__ != button_:
                self.deactivate_btn(button__)

    def deactivate_btn(self, button_: QPushButton):
        icon = QIcon()
        icon.addFile(self.menu_btns[button_]["icons"][0],
                     QSize(), QIcon.Mode.Normal, QIcon.State.Off)
        button_.setIcon(icon)
        button_.setStyleSheet(
            u"background: #transparent;"
            "color: #414141;"
        )

    def toggle_menu(self):
        menu = self.ui.menu_container
        menu.setMinimumWidth(170 if menu.minimumWidth() == 42 else 42)

    def hide_menu_buttons(self, buttons: list):
        for button in self.menu_btns:
            if button not in buttons:
                button.show()
            else:
                button.hide()

    def hide_dash_cards(self, cards: list):
        for card in self.display_cards:
            if card not in cards:
                card.show()
            else:
                card.hide()

    def setup_menu(self):
        for button in self.menu_btns:
            activate_func = functools.partial(self.activate_btn, button)
            button.clicked.connect(activate_func)
        self.menu_btn.clicked.connect(self.toggle_menu)
        self.home_btn.click()

    def build_menu(self, institution_type: str):
        self.edit_menu.clear()
        self.timetable_menu.clear()
        self.file_menu.clear()
        self.tools_menu.clear()

        if institution_type == Types.INSTITUTION_PRIMARY:
            buttons = [self.teachers_btn, self.lecturers_btn,
                       self.modules_btn, self.courses_btn,
                       self.blocks_btn]
            self.hide_menu_buttons(buttons)
            cards = [self.ui.modules_card, self.ui.lecturers_card,
                     self.ui.teachers_card, self.ui.blocks_card, self.ui.courses_card]
            self.hide_dash_cards(cards)
            self.ui.add_break_btn.show()

        elif institution_type == Types.INSTITUTION_SECONDARY:
            buttons = [self.lecturers_btn, self.modules_btn,
                       self.courses_btn]
            self.hide_menu_buttons(buttons)
            cards = [self.ui.courses_card, self.ui.modules_card,
                     self.ui.lecturers_card]
            self.hide_dash_cards(cards)
            self.ui.add_break_btn.show()

        elif institution_type == Types.INSTITUTION_COLLEGE:
            buttons = [self.teachers_btn, self.subjects_btn,
                       self.classes_btn, self.blocks_btn]
            self.hide_menu_buttons(buttons)
            cards = [self.ui.subjects_card, self.ui.classes_card,
                     self.ui.teachers_card, self.ui.blocks_card]
            self.hide_dash_cards(cards)
            self.ui.add_break_btn.hide()

    def set_saved(self, status: bool):
        self.save_indicator.setVisible(not status)

    def set_generated(self, status: bool):
        self.generate_indicator.setVisible(not status)

    """
    ======================================================
            Operations
            20/12/2025
    ======================================================
    """

    # def populate_projects(self):
    #     self.projects_list.clear()
    #     icon = QIcon()
    #     icon.addFile(u":/icons/icons/classes.svg", QSize(), QIcon.Mode.Normal, QIcon.State.Off)
    #     projects = Utils().get_projects()
    #     for project in projects:
    #         self.projects_list.addItem(icon, project)


class Boarding(QMainWindow):
    create = Signal(tuple)

    def __init__(self):
        super().__init__()
        self.ui = ui_boarding.Ui_Boarding()
        self.ui.setupUi(self)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.institution = None

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(30)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 80))
        self.ui.main_container.setGraphicsEffect(shadow)

        self.primary_school_btn = self.ui.primary_btn
        self.secondary_school_btn = self.ui.secondary_btn
        self.tertiary_school_btn = self.ui.tertiary_btn
        self.group = [self.primary_school_btn, self.secondary_school_btn,
                      self.tertiary_school_btn]
        self.ticks = [self.ui.tick_3, self.ui.tick_2, self.ui.tick]

        self.next_btn = self.ui.next_btn
        self.back_btn = self.ui.back_btn
        self.create_btn = self.ui.create_btn
        self.close_btn = self.ui.close_btn

        self.institution_name = self.ui.intitution_name_txt
        self.project_name = self.ui.timetable_name_txt
        self.daily_slots = self.ui.slots_per_day_txt
        self.days_per_cycle = self.ui.slots_per_cycle_txt
        self.slot_length = self.ui.slot_length_txt

        self.primary_school_btn.clicked.connect(self.mark_selection)
        self.secondary_school_btn.clicked.connect(self.mark_selection)
        self.tertiary_school_btn.clicked.connect(self.mark_selection)
        self.close_btn.clicked.connect(self.close)
        self.next_btn.clicked.connect(self.set_next)
        self.back_btn.clicked.connect(self.set_back)
        self.create_btn.clicked.connect(self.get_data)

        self.highlight(-1)

    def highlight(self, index):
        for tick in self.ticks:
            tick.hide()

        if index != -1:
            for button in self.group:
                button.setStyleSheet(f"QFrame#{button.objectName()}" + "{background: #F3F4F6;"
                                                                       "border: 1px solid #E3E4E6;"
                                                                       "border-radius: 10px;}")
            button = self.group[index]
            button.setStyleSheet(f"QFrame#{button.objectName()}" + "{background: #DBEAFE;"
                                                                   "border: 1px solid #E3E4E6;"
                                                                   "border-radius: 10px;}")
            self.ticks[index].show()

    def mark_selection(self, button: str):
        if button == "primary_btn":
            self.highlight(0)
            self.set_primary()
        if button == "secondary_btn":
            self.highlight(1)
            self.set_secondary()
        if button == "tertiary_btn":
            self.highlight(2)
            self.set_tertiary()

    def set_primary(self):
        self.institution = Types.INSTITUTION_PRIMARY

    def set_secondary(self):
        self.institution = Types.INSTITUTION_SECONDARY

    def set_tertiary(self):
        self.institution = Types.INSTITUTION_COLLEGE

    def set_next(self):
        if self.institution is not None:
            self.ui.container.setCurrentIndex(1)

    def set_back(self):
        self.ui.container.setCurrentIndex(0)

    def get_data(self):
        data = (
            self.institution, self.institution_name.text(),
            self.project_name.text(), int(self.daily_slots.text()),
            int(self.days_per_cycle.text()), int(self.slot_length.text())
        )
        institution, name, time_table_name, slots_per_day, days_per_cycle, slot_length = data

        if (institution and name and time_table_name and slots_per_day > 0
                and days_per_cycle > 0 and slot_length > 0):
            self.clear()
            self.create.emit(data)
        else:
            QMessageBox.warning(self, "Warning", "Please fill in all the fields.")

    def clear(self):
        self.institution = None
        self.institution_name.clear()
        self.project_name.clear()
        self.days_per_cycle.setValue(0)
        self.daily_slots.setValue(0)
        self.slot_length.setValue(0)


class StartUp(QMainWindow):
    project_open = Signal(str, str)
    project_create = Signal(str, str)

    def __init__(self):
        super().__init__()
        self.ui = ui_startup.Ui_StartUp()
        self.ui.setupUi(self)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.boarding = Boarding()
        self.central_cnt = self.ui.central_container

        self.prefs_btn = self.ui.prefs_btn
        self.web_btn = self.ui.online_btn
        self.close_btn = self.ui.close_btn
        self.minimize_btn = self.ui.minimize_btn
        self.quit_btn = self.ui.quit_btn

        self.search_txt = self.ui.projects_search_txt
        self.projects_table = self.ui.recent_layout
        self.create_project_btn = self.ui.create_new_project_btn
        self.open_project_btn = self.ui.open_project_btn
        self.import_project_btn = self.ui.import_project_btn
        self.recent_refresh_btn = self.ui.recent_projects_refresh_btn
        self.recent_clear_btn = self.ui.recents_clear_btn

        self.close_btn.clicked.connect(self.close)
        self.quit_btn.clicked.connect(self.close)
        self.minimize_btn.clicked.connect(self.showMinimized)
        self.prefs_btn.clicked.connect(self.show_prefs)
        self.web_btn.clicked.connect(self.visit_web)
        self.search_txt.textChanged.connect(self.search_recent)
        self.create_project_btn.clicked.connect(self.boarding.show)
        self.recent_refresh_btn.clicked.connect(self.populate_recent)
        self.recent_clear_btn.clicked.connect(self.recent_clear)

        self.populate_recent()

    def show_prefs(self):
        pass

    def visit_web(self):
        pass

    def populate_recent(self):
        try:
            self.clear_recent_layout()
            recent = Settings().get_recent_files()

            if not len(recent) > 0:
                self.ui.recent_status_txt.show()
            else:
                self.ui.recent_status_txt.hide()
                for name, path in recent:
                    self.show_recent(name, path)
        except Exception as e:
            logger.critical(str(e))

    def manage_recent(self, name: str, action: str):
        ...

    def clear_recent_layout(self):
        for i in range(self.projects_table.count() - 1):
            w = self.projects_table.takeAt(1).widget()
            w.setParent(None)

    def show_recent(self, name: str, path: str):
        self.projects_table.addWidget(RecentItem(
            self.ui.recents_container, name, path, callback=self.manage_recent
        ))

    def search_recent(self):
        recent = []
        search_text = self.search_txt.text()
        self.clear_recent_layout()

        if len(recent) == 0:
            self.projects_table.hide()
            self.ui.recent_status_txt.show()
        else:

            if search_text:
                pass
            else:
                self.projects_table.show()
                self.ui.recent_status_txt.hide()

                for name, path in recent:
                    self.show_recent(name, path)

    def recent_clear(self):
        Settings().clear_recent()
        self.populate_recent()
